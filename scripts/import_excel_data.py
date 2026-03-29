import os
import sys
import json
import re
from pathlib import Path

# 添加项目根目录到 Python 路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

# 导入你的模型
from backend.app.models import Base, Competition, Internship, PolicyLink, University
from backend.app.database import DATABASE_URL


class ExcelDataImporter:
    """Excel 数据导入器 - 修复版"""

    # 竞赛区域配置
    COMPETITION_CONFIG = {
        "Medicine": {"col_idx": 2, "col_letter": "B", "field_name": "Medicine"},
        "Engineering": {"col_idx": 4, "col_letter": "D", "field_name": "Engineering"},
        "Computer Sci/IT": {"col_idx": 6, "col_letter": "F", "field_name": "Computer Science"},
        "Business & Econ": {"col_idx": 8, "col_letter": "H", "field_name": "Business & Economics"},
        "Language & Culture": {"col_idx": 10, "col_letter": "J", "field_name": "Language & Culture"},
    }

    # 竞赛数据行范围
    COMPETITION_ROWS = [17, 18, 19, 20, 21]

    # 实习区域配置
    INTERNSHIP_CONFIG = {
        "Medicine": {"col_idx": 2, "col_letter": "B", "field_name": "Medicine"},
        "Engineering": {"col_idx": 4, "col_letter": "D", "field_name": "Engineering"},
        "Computer Sci/IT": {"col_idx": 6, "col_letter": "F", "field_name": "Computer Science"},
        "Business & Econ": {"col_idx": 8, "col_letter": "H", "field_name": "Business & Economics"},
        "Language & Culture": {"col_idx": 10, "col_letter": "J", "field_name": "Language & Culture"},
    }

    # 实习数据行范围
    INTERNSHIP_ROWS = [30, 31, 32, 33, 34]

    def detect_rows(self, ws, col_idx, pattern, start_row=10, end_row=80):
        rows = []
        rx = re.compile(pattern, flags=re.IGNORECASE)
        for r in range(start_row, end_row + 1):
            cell = ws.cell(r, col_idx)
            val = cell.value
            if isinstance(val, str) and rx.search(val.strip()):
                rows.append(r)
        return rows

    def __init__(self, excel_path: str, db_url: str = None):
        self.excel_path = Path(excel_path)
        self.db_url = db_url or DATABASE_URL
        self.engine = create_engine(self.db_url)
        self.SessionLocal = sessionmaker(bind=self.engine)

    def get_db(self):
        db = self.SessionLocal()
        try:
            yield db
        finally:
            db.close()

    def extract_hyperlink(self, cell):
        try:
            if hasattr(cell, "hyperlink") and cell.hyperlink:
                return cell.hyperlink.target
        except Exception:
            pass
        return None

    def clean_competition_name(self, raw_value):
        if not raw_value:
            return None
        name = str(raw_value).strip()
        patterns = [
            r"^\d+\.\s*",
            r"^Competition\s*\d+:\s*",
            r"^\d+\)\s*",
            r"^-\s*",
        ]
        for pattern in patterns:
            name = re.sub(pattern, "", name, flags=re.IGNORECASE)
        if name and name[0].isdigit() and ". " in name[:5]:
            parts = name.split(". ", 1)
            if len(parts) > 1:
                name = parts[1]
        return name.strip()

    def parse_competitions(self, ws):
        competitions = []
        print("\n📊 Parsing competitions...")
        print("   Column mapping:")
        auto_rows = self.detect_rows(ws, 2, r"^\s*Competition\s*\d+", start_row=10, end_row=80)
        rows_to_use = auto_rows if auto_rows else self.COMPETITION_ROWS
        for field_key, config in self.COMPETITION_CONFIG.items():
            col_idx = config["col_idx"]
            field_name = config["field_name"]
            col_letter = config["col_letter"]
            print(f"     {col_letter} ({col_idx}) -> {field_name}")
            for row_idx in rows_to_use:
                cell = ws.cell(row_idx, col_idx)
                raw_value = cell.value
                if raw_value:
                    comp_name = self.clean_competition_name(raw_value)
                    if comp_name:
                        skip_names = {
                            "internships/programmes",
                            "fields",
                            "university",
                            "city name",
                            "int student policy",
                            "university environment",
                            "characteristic expectations",
                        }
                        if comp_name.strip().lower() in skip_names or re.match(r"^intern/prog", comp_name.strip(), flags=re.IGNORECASE):
                            continue
                        comp_id = f"comp_{field_key.lower().replace(' ', '_')}_{row_idx}_{col_idx}"
                        link = self.extract_hyperlink(cell)
                        competition = {
                            "ext_id": comp_id,
                            "name": comp_name,
                            "fields_offered": json.dumps([field_name]),
                            "city": None,
                            "level": "International",
                            "link": link,
                            "deadline": None,
                            "description": f"{field_name} competition - {comp_name}",
                        }
                        competitions.append(competition)
                        print(f"     ✓ [{field_name}] Row {row_idx}: {comp_name[:50]}")
                else:
                    print(f"     ⚠ [{field_name}] Row {row_idx}: empty")
        return competitions

    def clean_internship_name(self, raw_value):
        if not raw_value:
            return None
        name = str(raw_value).strip()
        patterns = [
            r"^\d+\.\s*",
            r"^Intern/prog\s*\d+:\s*",
            r"^\d+\)\s*",
            r"^-\s*",
        ]
        for pattern in patterns:
            name = re.sub(pattern, "", name, flags=re.IGNORECASE)
        if name.startswith("=HYPERLINK"):
            match = re.search(r'=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)', name)
            if match:
                return match.group(2).strip()
        return name.strip()

    def extract_company_from_name(self, name: str) -> str:
        if not name:
            return "Various"
        companies = [
            "Google",
            "Microsoft",
            "Amazon",
            "Stanford",
            "Girls Who Code",
            "Bank of America",
            "Harvard",
            "Wharton",
            "JA",
            "FBLA",
            "CIEE",
            "NSLI-Y",
            "AFS",
            "YFU",
            "Fulbright",
            "Rolls-Royce",
            "NTU",
            "Imperial",
            "United Planet",
            "AIMI",
            "Smith College",
        ]
        for company in companies:
            if company.lower() in name.lower():
                return company
        first_word = name.split()[0] if name.split() else "Various"
        return first_word

    def parse_internships(self, ws):
        internships = []
        print("\n💼 Parsing internships...")
        print("   Column mapping:")
        auto_rows = self.detect_rows(ws, 2, r"^\s*Intern/prog\s*\d+", start_row=10, end_row=80)
        rows_to_use = auto_rows if auto_rows else self.INTERNSHIP_ROWS
        for field_key, config in self.INTERNSHIP_CONFIG.items():
            col_idx = config["col_idx"]
            field_name = config["field_name"]
            col_letter = config["col_letter"]
            print(f"     {col_letter} ({col_idx}) -> {field_name}")
            for row_idx in rows_to_use:
                cell = ws.cell(row_idx, col_idx)
                raw_value = cell.value
                if raw_value:
                    intern_name = self.clean_internship_name(raw_value)
                    link = self.extract_hyperlink(cell)
                    if not link and isinstance(raw_value, str) and raw_value.startswith("=HYPERLINK"):
                        match = re.search(r'=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)', raw_value)
                        if match:
                            link = match.group(1)
                            intern_name = match.group(2)
                    if intern_name:
                        skip_names = {
                            "internships/programmes",
                            "fields",
                            "university",
                            "city name",
                            "int student policy",
                            "university environment",
                            "characteristic expectations",
                        }
                        if intern_name.strip().lower() in skip_names:
                            continue
                        intern_id = f"intern_{field_key.lower().replace(' ', '_')}_{row_idx}_{col_idx}"
                        internship = {
                            "ext_id": intern_id,
                            "company": self.extract_company_from_name(intern_name),
                            "name": intern_name,
                            "fields_offered": json.dumps([field_name]),
                            "city": None,
                            "link": link,
                            "deadline": None,
                            "description": f"{field_name} internship opportunity",
                            "requirements": json.dumps({}),
                        }
                        internships.append(internship)
                        print(f"     ✓ [{field_name}] Row {row_idx}: {intern_name[:50]} {'🔗' if link else ''}")
                else:
                    print(f"     ⚠ [{field_name}] Row {row_idx}: empty")
        return internships

    def debug_excel_structure(self, ws):
        print("\n🔍 DEBUG: Excel structure preview")
        print("   Showing rows 15-40, columns B-J (2-10)")
        print("   " + "-" * 80)
        for row in range(15, 41):
            row_data = []
            for col in [2, 4, 6, 8, 10]:
                cell = ws.cell(row, col)
                if cell.value:
                    val = str(cell.value)[:40]
                    if val.startswith("=HYPERLINK"):
                        val = val[:30] + "..."
                    row_data.append(f"{get_column_letter(col)}{row}: {val}")
            if row_data:
                print(f"   Row {row}: {', '.join(row_data)}")
            elif row in [16, 17, 18, 19, 20, 21, 27, 28, 29, 30, 31, 32, 33, 34]:
                print(f"   Row {row}: (empty)")
        print("   " + "-" * 80)

    def clear_existing_data(self, db: Session, clear_competitions=True, clear_internships=True, clear_policies=True):
        if clear_competitions:
            count = db.query(Competition).delete()
            print(f"   Cleared {count} competitions")
        if clear_internships:
            count = db.query(Internship).delete()
            print(f"   Cleared {count} internships")
        if clear_policies:
            count = db.query(PolicyLink).delete()
            print(f"   Cleared {count} policies")
        db.commit()

    def import_all(self, clear_first: bool = False, debug: bool = False):
        print(f"\n📁 Loading Excel file: {self.excel_path}")
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
        wb = load_workbook(self.excel_path, data_only=False)
        ws = wb["Sheet1"]
        if debug:
            self.debug_excel_structure(ws)
        db = self.SessionLocal()
        try:
            if clear_first:
                print("\n🗑️  Clearing existing data...")
                self.clear_existing_data(db)
            competitions = self.parse_competitions(ws)
            print(f"\n   ✅ Found {len(competitions)} competitions")
            for comp_data in competitions:
                existing = db.query(Competition).filter(Competition.name == comp_data["name"]).first()
                if existing:
                    for key, value in comp_data.items():
                        if value is not None:
                            setattr(existing, key, value)
                else:
                    comp = Competition(**comp_data)
                    db.add(comp)
            db.commit()
            print(f"   ✅ Imported {len(competitions)} competitions")
            internships = self.parse_internships(ws)
            print(f"\n   ✅ Found {len(internships)} internships")
            for intern_data in internships:
                existing = db.query(Internship).filter(Internship.name == intern_data["name"]).first()
                if existing:
                    for key, value in intern_data.items():
                        if value is not None:
                            setattr(existing, key, value)
                else:
                    intern = Internship(**intern_data)
                    db.add(intern)
            db.commit()
            print(f"   ✅ Imported {len(internships)} internships")
            print("\n" + "=" * 50)
            print("✅ All data imported successfully!")
            print("=" * 50)
            print("\n📊 Database Statistics:")
            print(f"   - Competitions: {db.query(Competition).count()}")
            print(f"   - Internships: {db.query(Internship).count()}")
            print("\n📋 Sample imported competitions:")
            for comp in db.query(Competition).limit(5):
                print(f"   - {comp.name} | Fields: {comp.fields_offered} | Link: {'Yes' if comp.link else 'No'}")
            print("\n📋 Sample imported internships:")
            for intern in db.query(Internship).limit(5):
                print(f"   - {intern.name} | Company: {intern.company} | Link: {'Yes' if intern.link else 'No'}")
        except Exception as e:
            print(f"\n❌ Error during import: {e}")
            db.rollback()
            raise
        finally:
            db.close()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Import Excel data to database")
    parser.add_argument("--excel", type=str, default="knowledge/Hackathon_Data sheet (1).xlsx", help="Excel file path")
    parser.add_argument("--clear", action="store_true", help="Clear existing data before import")
    parser.add_argument("--debug", action="store_true", help="Print debug info about Excel structure")
    parser.add_argument("--db", type=str, default=None, help="Database URL (default from config)")
    args = parser.parse_args()
    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"❌ Excel file not found: {excel_path}")
        print("Please provide the correct path using --excel parameter")
        sys.exit(1)
    importer = ExcelDataImporter(str(excel_path), args.db)
    importer.import_all(clear_first=args.clear, debug=args.debug)
    print("\n✨ Done!")

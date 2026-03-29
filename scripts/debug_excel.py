import sys
from pathlib import Path
from openpyxl import load_workbook

excel_path = "Hackathon_Data sheet (1).xlsx"
if not Path(excel_path).exists():
    alt = Path("knowledge") / "Hackathon_Data sheet (1).xlsx"
    if alt.exists():
        excel_path = str(alt)
    else:
        print(f"File not found: {excel_path}")
        sys.exit(1)

wb = load_workbook(excel_path, data_only=False)
ws = wb["Sheet1"]

print("=" * 80)
print("EXCEL STRUCTURE DEBUG")
print("=" * 80)

print("\n📊 COMPETITION MATCHING SECTION (Rows 15-25):")
print("-" * 80)
for row in range(15, 26):
    row_data = []
    for col in [2, 4, 6, 8, 10]:
        cell = ws.cell(row, col)
        if cell.value:
            val = str(cell.value)
            has_link = hasattr(cell, "hyperlink") and cell.hyperlink
            link_target = cell.hyperlink.target if has_link else None
            display = f"{val[:50]}"
            if link_target:
                display += f" [LINK: {link_target[:50]}]"
            row_data.append(display)
        else:
            row_data.append("(empty)")
    if any(r != "(empty)" for r in row_data):
        print(f"Row {row}: B={row_data[0]}, D={row_data[1]}, F={row_data[2]}, H={row_data[3]}, J={row_data[4]}")

print("\n\n💼 INTERNSHIPS SECTION (Rows 27-38):")
print("-" * 80)
for row in range(27, 39):
    row_data = []
    for col in [2, 4, 6, 8, 10]:
        cell = ws.cell(row, col)
        if cell.value:
            val = str(cell.value)
            has_link = hasattr(cell, "hyperlink") and cell.hyperlink
            link_target = cell.hyperlink.target if has_link else None
            display = f"{val[:50]}"
            if link_target:
                display += f" [LINK: {link_target[:50]}]"
            row_data.append(display)
        else:
            row_data.append("(empty)")
    if any(r != "(empty)" for r in row_data):
        print(f"Row {row}: B={row_data[0]}, D={row_data[1]}, F={row_data[2]}, H={row_data[3]}, J={row_data[4]}")

print("\n" + "=" * 80)
